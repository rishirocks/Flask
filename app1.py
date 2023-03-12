from flask import Flask,render_template,request
from openpyxl import load_workbook


app = Flask(__name__)
wb = load_workbook('Marks.xlsx')
sheet = wb.active
row = sheet.max_row+1
@app.route("/",methods=["POST","GET"])

def submit_marks():
    global row

    IA_one = ''
    IA_two = ''
    Reg_no = ''
    Assign = ''
    if request.method == "POST":
       if request.form.get('Reg_no'):
         Reg_no = request.form.get('Reg_no')
       else:
         Reg_no = None

       if request.form.get('IA_1'):
         IA_1 = int(request.form.get('IA_1'))
         IA_one = float((IA_1 * 30) / 100)
       else:
         IA_one = None

       if request.form.get('IA_2'):
         IA_2 = int(request.form.get('IA_2'))
         IA_two = float((IA_2 * 30) / 100)
       else:
         IA_two =None

       if request.form.get('Assign'):
         Assign = int(request.form.get('Assign'))
       else:
         Assign = None

       if sheet.cell(row=1,column=1).value is None and sheet.cell(row=1,column=2).value is None and sheet.cell(row=1,column=3).value is None and sheet.cell(row=1,column=4).value is None:
          sheet.cell(row=1,column=1,value='Reg_no')
          sheet.cell(row=1, column=2, value='IA_one')
          sheet.cell(row=1, column=3, value='IA_two')
          sheet.cell(row=1, column=4, value='Assig')

       for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == Reg_no:
                sheet.cell(row=row, column=2).value = IA_one
                sheet.cell(row=row, column=3).value = IA_two
                sheet.cell(row=row, column=4).value = Assign
                break
            elif sheet.cell(row=row,column=2).value is None:
               sheet.cell(row=row,column=2).value = IA_one
            elif sheet.cell(row=row,column=3).value is None:
               sheet.cell(row=row,colummn=3).value = IA_two
            elif sheet.cell(row=row,column=4).value is None:
               sheet.cell(row=row,column=4).value = Assign
       else:
          data = [Reg_no,IA_one,IA_two,Assign]
          sheet.append(data)
       wb.save('Marks.xlsx')
       row +=1
    return render_template("home.html",Reg_no=Reg_no,IA_one=IA_one,IA_two=IA_two,Assign=Assign)

if __name__=="__main__":
    app.run(debug=True)
